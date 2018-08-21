"""记录已爬取过的数据"""
import os
import openpyxl
import json
from util.configure import column

path_recorder = os.path.join(os.getcwd(), 'record')
path_excel = os.path.join(os.getcwd(), 'result.xlsx')
wb = openpyxl.load_workbook(path_excel)
sheet = wb.active


def recorder(path):
    record_dict = {}

    for i in range(2, sheet.max_row + 1):
        if sheet[column["citedby"] + str(i)].value is not None:
            record_dict[i] = True
        else:
            record_dict[i] = False

    with open(path_recorder, 'w', encoding='utf-8') as f:
        f.write(json.dumps(record_dict))


if __name__ == '__main__':
    recorder(path_excel)
    with open(path_recorder, 'r', encoding='utf-8') as f:
        dict_recorder = json.loads(f.read())
        count = 0
        for item in dict_recorder.values():
            if item is True:
                count += 1
        print(count, '/', str(sheet.max_row - 1))
