import os
import openpyxl
import scholar as scholarly
import multiprocessing
from SpiderKit.ip_pool_foreign import IPProvider
import time
import json
import sys

column = {
    "expert": "A",
    "affiliation": "B",
    "interests": "C",
    "name": "D",
    'email':'E',
    "citedby": "F",
    "hindex": "G",
    "hindex5y": "H",
    "i10index": "I",
    "i10index5y": "J",
    "url_picture": "K",
}
path_excel = os.path.join(os.getcwd(), 'result.xlsx')
path_recorder = os.path.join(os.getcwd(), 'record')
path_result_file = os.path.join(os.getcwd(), 'result')

# 读取文件
reference = openpyxl.load_workbook(path_excel)
sheet = reference.active
with open(path_recorder, 'r', encoding='utf-8') as f:
    # 记录结果字典：True - 结果已存在， False - 不存在
    recoder_list = json.loads(f.read())


def complement(lock, lower, upper, batch=5):
    """
    专家信息补齐
    :param lock: 多进程锁
    :param lower: 处理学者id下界
    :param upper: 处理学者id上界
    :param batch: 分批次处理 批次大小
    :return:
    """
    print(lower, "~", upper)
    while lower < upper:
        list_result = []

        # 获取代理ip，记录提取的时间
        ipprovider = IPProvider()
        start = time.time()
        proxy = ipprovider.get_ip()

        # 是否更换ip
        _isIPNeedChange = False
        for i in range(lower, min(lower + batch, upper + 1)):
            if i > sheet.max_row:
                break
            if recoder_list.setdefault(str(i), False):
                continue

            # 如果ip已经用了30秒，或者需要更换ip，则更换ip
            if time.time() - start >= 30 or _isIPNeedChange:
                start = time.time()
                proxy = ipprovider.get_ip()
                _isIPNeedChange = False
            expert = sheet[column["expert"] + str(i)].value
            expert = expert if expert is not None else ''

            author = None
            none_author_tuple = (str(i), '', '', '', -1, -1, -1, -1, -1, '')
            # 最多重试三次
            max_tries = 3
            while author is None and max_tries > 0:
                try:
                    author = next(scholarly.search_author(expert, proxy)).fill(proxy)
                # 如果谷歌学术中不存在该学者的信息，则记录默认值
                except StopIteration:
                    print('No professor named', expert, i)
                    list_result.append((none_author_tuple))
                    break
                # 如果出现网络错误，则请求更换ip
                except Exception as e:
                    _isIPNeedChange = True
                    print(e, expert, i)
                    # 时间惩罚
                    time.sleep(60)
                    max_tries -= 1

            if author is None:
                list_result.append(none_author_tuple)
                continue

            name = author.name
            affiliation = author.affiliation
            email = author.email
            citedby = author.citedby
            hindex = author.hindex
            hindex5y = author.hindex5y
            i10index = author.i10index
            i10index5y = author.i10index5y
            url_picture = author.url_picture

            # 列表记录爬取结果
            result = (str(i), name, affiliation, email, citedby, hindex, hindex5y, i10index, i10index5y, url_picture)
            list_result.append(result)
            print(result)

        lock.acquire()
        # 每批次执行一次：记录结果
        try:
            with open(path_result_file, 'a', encoding='utf-8') as file:
                for r in list_result:
                    file.write(json.dumps(r) + '\n')

        finally:
            lock.release()

        lower = min(upper, lower + batch)


if __name__ == '__main__':
    with open(path_result_file, 'w', encoding='utf-8') as f:
        pass

    """
    多进程执行
    :param begin_no: 爬取开始id
    :param counts: 这一次需要爬取多少数据
    :param num_of_processing: 进程个数
    :param batch: 分批次处理 批次大小
    
    执行示例：python complement.py [begin_no] [counts] [num_of_processing]
    """
    begin_no = int(sys.argv[1])
    counts = int(sys.argv[2])
    num_of_processing = 4 if len(sys.argv) < 4 else int(sys.argv[3])
    have_done = 0

    if begin_no + counts > sheet.max_row:
        counts = sheet.max_row - begin_no + 1

    quarter = counts // num_of_processing
    lock = multiprocessing.Lock()
    arg_list = [
        (lock, have_done + begin_no, begin_no + quarter),
        (lock, have_done + begin_no + quarter + 1, begin_no + quarter * 2),
        (lock, have_done + begin_no + quarter * 2 + 1, begin_no + quarter * 3),
        (lock, have_done + begin_no + quarter * 3 + 1, begin_no + quarter * 4),
        (lock, have_done + begin_no + quarter * 4 + 1, begin_no + quarter * 5),
        (lock, have_done + begin_no + quarter * 5 + 1, begin_no + quarter * 6),
    ]

    for j in range(1, num_of_processing + 1):
        process = multiprocessing.Process(target=complement, args=arg_list[j - 1])
        process.start()
        time.sleep(30)
