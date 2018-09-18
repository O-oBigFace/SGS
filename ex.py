import os
import openpyxl
import scholar as scholarly
import multiprocessing
from SpiderKit.ip_pool_foreign import IPProvider
import time
import json
import sys
import util.configure
import Google_complement

column = util.configure.column

path_excel = os.path.join(os.getcwd(), 'result.xlsx')
path_result_file = os.path.join(os.getcwd(), 'result_ex')

# 读取文件
reference = openpyxl.load_workbook(path_excel)
sheet = reference.active


def complement(lock, lower, upper, batch=5):
    """
    专家信息补齐
    :param lock: 多进程锁
    :param lower: 处理学者id下界
    :param upper: 处理学者id上界
    :param batch: 分批次处理 批次大小
    :return:
    """
    print(lower, "~", upper)
    while lower < upper:
        list_result = []

        proxy = ''
        for i in range(lower, min(lower + batch, upper + 1)):
            if i > sheet.max_row:
                break

            name = sheet[column["name"] + str(i)].value
            _affiliation = sheet[column['affiliation'] + str(i)].value

            if len(_affiliation) < 5:
                author = next(scholarly.search_author(name, proxy)).fill(proxy)
                _affiliation = author.affiliation

            keywords = name + ' and ' + _affiliation
            _email, phone = Google_complement.get_email_and_phone(keywords)
            address = Google_complement.get_address(_affiliation)
            country = Google_complement.get_country(_affiliation)

            # 列表记录爬取结果
            result = (str(i),
                      _affiliation,
                      _email,
                      phone,
                      address,
                      country
                      )
            list_result.append(result)
            print(result)

        lock.acquire()
        # 每批次执行一次：记录结果
        try:
            with open(path_result_file, 'a', encoding='utf-8') as file:
                for r in list_result:
                    file.write(json.dumps(r) + '\n')
        finally:
            lock.release()

        lower = min(upper, lower + batch)


def input():
    with open(path_result_file, 'r', encoding='utf-8') as f:
        for line in f.readlines():
            item = json.loads(line.strip())
            print(item)
            sheet[column["affiliation"] + item[0]] = item[1]
            sheet[column["_email"] + item[0]] = item[2]
            sheet[column["phone"] + item[0]] = item[3]
            sheet[column["address"] + item[0]] = item[4]
            sheet[column["country"] + item[0]] = item[5]
    reference.save(path_excel)

#
# if __name__ == '__main__':
#     # 清空result文件，防止其过大
#     with open(path_result_file, 'w', encoding='utf-8') as f:
#         pass
#
#     """
#     多进程执行
#     :param begin_no: 爬取开始id
#     :param counts: 这一次需要爬取多少数据
#     :param num_of_processing: 进程个数
#     :param batch: 分批次处理 批次大小
#
#     执行示例：python complement.py [begin_no] [counts] [num_of_processing]
#     """
#     begin_no = int(sys.argv[1])
#     counts = int(sys.argv[2])
#
#     num_of_processing = 4 if len(sys.argv) < 4 else int(sys.argv[3])
#     have_done = 0
#
#     if begin_no + counts > sheet.max_row:
#         counts = sheet.max_row - begin_no + 1
#
#     quarter = counts // num_of_processing
#     lock = multiprocessing.Lock()
#     arg_list = [
#         (lock, have_done + begin_no, begin_no + quarter),
#         (lock, have_done + begin_no + quarter + 1, begin_no + quarter * 2),
#         (lock, have_done + begin_no + quarter * 2 + 1, begin_no + quarter * 3),
#         (lock, have_done + begin_no + quarter * 3 + 1, begin_no + quarter * 4),
#         (lock, have_done + begin_no + quarter * 4 + 1, begin_no + quarter * 5),
#         (lock, have_done + begin_no + quarter * 5 + 1, begin_no + quarter * 6),
#         (lock, have_done + begin_no + quarter * 6 + 1, begin_no + quarter * 7),
#         (lock, have_done + begin_no + quarter * 7 + 1, begin_no + quarter * 8),
#     ]
#
#     for j in range(1, num_of_processing + 1):
#         process = multiprocessing.Process(target=complement, args=arg_list[j - 1])
#         process.start()
#         time.sleep(30)


if __name__ == '__main__':
    input()